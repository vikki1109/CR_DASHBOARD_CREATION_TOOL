import glob
import os

import time
import datetime
import pandas as pd
from datetime import datetime, timedelta
import tkinter as tk
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles import PatternFill,Border, Side
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import ttk
from tkcalendar import Calendar
import xlwings as xw
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from webdriver_manager.microsoft import EdgeChromiumDriverManager

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.keys import Keys

 
 
def update_start_date(event):
    start_date_label.config(text=f"Start Date: {cal_start.get_date()} {start_time_entry.get()}")
 
 
def update_end_date(event):
    end_date_label.config(text=f"End Date: {cal_end.get_date()} {end_time_entry.get()}")
 
 
def submit_dates():
    global start_date1, end_date1, vendor
    _start_date = cal_start.get_date()
    _end_date = cal_end.get_date()
    start_date1 = f"{_start_date} {start_time_entry.get()}".strip()
    end_date1 = f"{_end_date} {end_time_entry.get()}".strip()
    vendor = vendor_combobox.get()  # Get the selected vendor
    result_label.config(text=f"Start Date: {start_date1}\nEnd Date: {end_date1}\nVendor: {vendor}")
 
root = tk.Tk()
root.title("Date and Time Range Selector")
root.geometry("400x650")
 
style = ttk.Style()
root.configure(background='light blue')
style.configure("TLabel", background="light blue", foreground="black", font=("Arial", 12, "bold"))
style.configure("TButton", background="white", foreground="black", font=("Arial", 12), relief="flat")
 
 
start_date_label = ttk.Label(root, text="Start Date")
start_date_label.pack()
 
 
cal_start = Calendar(root, selectmode="day", year=2023, month=10, day=17, date_pattern="dd/mm/yyyy")
cal_start.pack(pady=7)
cal_start.bind("<<CalendarSelected>>", update_start_date)
 
 
end_date_label = ttk.Label(root, text="End Date")
end_date_label.pack()
 
 
cal_end = Calendar(root, selectmode="day", year=2023, month=10, day=17, date_pattern="dd/mm/yyyy")
cal_end.pack(pady=7)
cal_end.bind("<<CalendarSelected>>", update_end_date)
 
 
def validate_time_format(P):
    return P == "" or (":" in P and all(part.isdigit() for part in P.split(":")))
vcmd = (root.register(validate_time_format), "%P")
 
 
start_time_label = ttk.Label(root, text="Start Time (HH:MM:SS): ")
start_time_label.pack()
start_time_entry = ttk.Entry(root, validate="key", validatecommand=vcmd)
start_time_entry.insert(0, "08:00:00")  # Default start time
start_time_entry.pack()
 
end_time_label = ttk.Label(root, text="End Time (HH:MM:SS): ")
end_time_label.pack()
end_time_entry = ttk.Entry(root, validate="key", validatecommand=vcmd)
end_time_entry.insert(0, "08:00:00")  # Default start time
end_time_entry.pack()
 
 
vendor_label = ttk.Label(root, text="Select Vendor:")
vendor_label.pack()
vendor_combobox = ttk.Combobox(root, values=["Nokia", "ZTE"])
vendor_combobox.pack()
 
 
submit_button = ttk.Button(root, text="Submit", command=lambda: [submit_dates(), root.destroy()])
submit_button.pack(pady=10)
 
 
result_label = ttk.Label(root, text="")
result_label.pack()
root.mainloop()
script_directory = os.path.dirname(os.path.abspath(__file__))
# Define the relative path to the 'data' folder
files_path = os.path.join(script_directory, 'data')
options = webdriver.EdgeOptions()
prefs={"download.default_directory":files_path}
options.add_experimental_option("prefs",prefs)
driver = webdriver.Chrome(service=Service(EdgeChromiumDriverManager().install()),options=options)
 
driver.maximize_window()
 
time.sleep(1)
 
driver.get('https://nextgentm-in.sdt.ericsson.net/arsys/forms/umt-ars-in/SHR%3ALandingConsole/Default+Administrator+View/?cacheid=c4ed3626')
 
while True:
    try:        
        if driver.find_element(By.XPATH,'*//label[text()="Company"]').text == 'Company':
            break
    except:
        pass
time.sleep(3)
try:
    driver.find_element(By.XPATH,'/html/body/div[1]/div[5]/div[2]/div/div/div[3]/fieldset/div/div/div/div/div[1]/fieldset/div/div[2]/fieldset/a[1]').click()
except:
    pass
 
driver.find_element(By.LINK_TEXT,'Smart Reporting').click()
driver.find_element(By.LINK_TEXT,'Smart Reporting Console').click()
driver.switch_to.window(driver.window_handles[1])
 
time.sleep(5)
while True:
    try:       
        if 'Client:' in driver.find_element(By.XPATH,'*//span[text()="Client: "]').text:
            break
    except:
        pass

time.sleep(3)
driver.find_element(By.XPATH,'/html/body/form/div[2]/div[3]/div/div/input').send_keys('Dashboard'+Keys.ENTER)
time.sleep(5)
a = ActionChains(driver)
a.move_to_element(driver.find_element(By.XPATH,'//div[text()="Dashboard"]')).double_click().perform()
while True:
    try:        
        if driver.find_element(By.XPATH,'//span[text()="Scheduled Start Date"]').text == 'Scheduled Start Date':
            break
    except:
        pass
time.sleep(3)
end_date = end_date1
start_date = start_date1
driver.find_element(By.XPATH,'/html/body/div[5]/div[2]/div[3]/div[1]/div/div/div[2]/div[6]/div/div[4]/div/div/div[1]/div/input').click()
time.sleep(2)
driver.find_element(By.XPATH,'/html/body/div[8]/div[2]/div[2]/div[1]/div[1]/input').send_keys(Keys.CONTROL+'a')
time.sleep(0.5)
driver.find_element(By.XPATH,'/html/body/div[8]/div[2]/div[2]/div[1]/div[1]/input').send_keys(start_date)
time.sleep(1)
driver.find_element(By.XPATH,'/html/body/div[8]/div[2]/div[2]/div[1]/div[2]/input').send_keys(Keys.CONTROL+'a')
time.sleep(0.5)
driver.find_element(By.XPATH,'/html/body/div[8]/div[2]/div[2]/div[1]/div[2]/input').send_keys(end_date)
time.sleep(1)
driver.find_element(By.XPATH,'/html/body/div[8]/div[2]/div[2]/div[2]/div[2]/table/tbody/tr/td/div/table/tbody/tr/td[2]').click()
time.sleep(3)
driver.find_element(By.XPATH,'/html/body/div[5]/div[2]/div[3]/div[1]/div/div/div[3]/div[1]/table/tbody/tr/td/div/table/tbody/tr/td[2]').click()
time.sleep(5)
driver.find_element(By.XPATH,'/html/body/div[5]/div[1]/div/div[1]/div[2]/table[2]/tbody/tr/td[1]').click()
time.sleep(3)
driver.find_element(By.LINK_TEXT,'Export to XLSX').click()
time.sleep(3)
driver.find_element(By.XPATH,'*//button/span[text()="Export"]').click()
time.sleep(5)
driver.close()
driver.switch_to.window(driver.window_handles[0])
time.sleep(3)
driver.find_element(By.XPATH,'*//div/div[text()="Logout"]').click()
time.sleep(2)
driver.close()
# Use glob to find Excel files in the 'data' folder
file_type = '*.xlsx'
files = glob.glob(os.path.join(files_path, file_type))
# Check if any files were found
if not files:
    print("No .xlsx files found in the directory.")
else:
    # Access the most recently created file in the list
    file_path = max(files, key=os.path.getctime)
    print(file_path)
df = pd.read_excel(file_path)
# Delete the first 3 rows
df = df.iloc[2:]
# Save the modified DataFrame back to a new Excel file
new_file_path = os.path.join(files_path,'Dashboard.xlsx')
df.to_excel(new_file_path, index=False)
columns_to_read = ['Change ID','Assignee Group','Impact','Change Request Status','Status Reason','Circle','Domain','Parent Category','Child Category','Summary','Scheduled Start Date','Scheduled End Date','No of Nodes','Execution Type','Closure Comment','Custom_Field10']
data = pd.read_excel(os.path.join(files_path,'Dashboard.xlsx'), usecols=columns_to_read, engine='openpyxl',skiprows=1)
date_column = 'Scheduled Start Date'
data['Scheduled Start Date'] = pd.to_datetime(data[date_column])
modified_dates = []
for index, row in data.iterrows():
    try:
        start_time3 = row['Scheduled Start Date']
        if start_time3.hour < 4:
            modified_date = (start_time3 - timedelta(days=1)).strftime('%d-%b')
        else:
            modified_date = start_time3.strftime('%d-%b')        
        modified_dates.append(modified_date)
    except AttributeError:
        modified_dates.append("Invalid Date")
       
data['Date'] = modified_dates
end_date = data[date_column].max()
start_date = end_date - timedelta(days=7)  # Last 1 week
filtered_data = data[(data[date_column] >= start_date) & (data[date_column] <= end_date)]
Change_column = 'Assignee Group'
filtered_data = filtered_data[(filtered_data[Change_column] == 'SRF-RAN CM Delhi')]
Change_column = 'Custom_Field10'
filtered_data = filtered_data[(filtered_data[Change_column] == 'Nokia') | (filtered_data[Change_column] == 'ZTE') | (filtered_data[Change_column] == 'Samsung')| (filtered_data[Change_column] == 'Ericsson')| (filtered_data[Change_column] == 'Huawei')]
Change_column = 'Change Request Status'
filtered_data = filtered_data[(filtered_data[Change_column] == 'Cancelled') | (filtered_data[Change_column] =='Closed') | (filtered_data[Change_column]=='Completed') | (filtered_data[Change_column]=='Implementation In Progress') | (filtered_data[Change_column] =='Scheduled')]
filtered_data.to_excel(os.path.join(script_directory, 'Automation data_Week-xx.xlsx'))
Change_column = 'Circle'
filtered_data = filtered_data[(filtered_data[Change_column] == 'BH') | (filtered_data[Change_column] =='GJ') | (filtered_data[Change_column]=='KL') | (filtered_data[Change_column]=='KK') | (filtered_data[Change_column]=='HR')   | (filtered_data[Change_column]=='AP')  | (filtered_data[Change_column]=='AS')  | (filtered_data[Change_column]=='CH')  | (filtered_data[Change_column]=='DL')  | (filtered_data[Change_column]=='HP')  | (filtered_data[Change_column]=='JH')  | (filtered_data[Change_column]=='JK')  | (filtered_data[Change_column]=='KO')  | (filtered_data[Change_column]=='MH')  | (filtered_data[Change_column]=='MP')  | (filtered_data[Change_column]=='MU')  | (filtered_data[Change_column]=='NE')  | (filtered_data[Change_column]=='OR')  | (filtered_data[Change_column]=='PB')  | (filtered_data[Change_column]=='RJ')  | (filtered_data[Change_column]=='TN')  | (filtered_data[Change_column]=='UPE') | (filtered_data[Change_column]=='UPW') | (filtered_data[Change_column]=='WB')]
print(files_path)
filtered_data = filtered_data.drop(['Assignee Group'],axis=1)
filtered_data.loc[filtered_data['Status Reason']=='Backed Out','Change Request Status'] = 'Reverted'
filtered_data.loc[filtered_data['Change Request Status']=='Closed','Change Request Status'] = 'Completed'
filtered_data = filtered_data.drop(['Status Reason'],axis=1)
column_to_split = 'No of Nodes'
delimiter = r'[/\\]'
# Split the column using the delimiter
split_data = filtered_data[column_to_split].str.split(delimiter, n=1, expand=True)

if split_data[1] is not  None:
    # Delimiter is present, assign 'Done Count' and 'Planned Count'
    filtered_data[['Done Count', 'Planned Count']] = split_data
    filtered_data['Done Count'] = pd.to_numeric(filtered_data['Done Count'], errors='coerce')
    filtered_data['Planned Count'] = pd.to_numeric(filtered_data['Planned Count'], errors='coerce')
else:
    # Delimiter is not present, assign the whole value to 'Planned Count' and 0 to 'Done Count'
    filtered_data['Planned Count'] = pd.to_numeric(filtered_data[column_to_split], errors='coerce')
    filtered_data['Done Count'] = 0
done_count_sum = filtered_data['Done Count'].sum()
planned_count_sum = filtered_data['Planned Count'].sum() 
print(f"Sum of 'Done Count': {done_count_sum}")
print(f"Sum of 'Planned Count': {planned_count_sum}")
filtered_data = filtered_data.drop(['No of Nodes'],axis=1)
filtered_data.loc[filtered_data['Change Request Status']=='Cancelled','Execution Type'] = 'Cancelled'
filtered_data.loc[filtered_data['Change Request Status']=='Reverted','Execution Type'] = 'Reverted'
filtered_data.to_excel(os.path.join(script_directory, 'Automation data_Week-xx.xlsx'))
df = pd.read_excel(os.path.join(script_directory, 'Automation data_Week-xx.xlsx'))
df.rename(columns={'Custom_Field10': 'Vendor'}, inplace=True)
df.to_excel(os.path.join(script_directory, 'Automation data_Week-xx.xlsx'))
df = pd.read_excel(os.path.join(script_directory, 'Automation data_Week-xx.xlsx'))
df['S No'] = range(1, len(df) + 1)
desired_column_order = ['S No','Date','Change ID', 'Impact', 'Circle', 'Domain', 'Parent Category', 'Child Category', 'Summary', 'Change Request Status', 'Planned Count', 'Done Count', 'Closure Comment', 'Vendor', 'Scheduled Start Date', 'Scheduled End Date', 'Execution Type']
df = df[desired_column_order]
df.to_excel(os.path.join(script_directory, 'Automation data_Week-xx.xlsx'),index=False)
from openpyxl.styles import Font, Alignment
workbook = openpyxl.load_workbook(os.path.join(script_directory, 'Automation data_Week-xx.xlsx'))
sheet_index = 0 
sheet_to_rename = workbook.worksheets[sheet_index]
sheet_name_to_rename = 'Sheet1'
sheet_to_rename = workbook[sheet_name_to_rename]
new_sheet_name = 'Raw Data'  
sheet_to_rename.title = new_sheet_name
new_sheet_name = 'Dashboard'  
new_sheet = workbook.create_sheet(title='Dashboard')
workbook.save(os.path.join(script_directory,'Automation data_Week-xx.xlsx'))
workbook = openpyxl.load_workbook(os.path.join(script_directory, 'Automation data_Week-xx.xlsx'))
sheet = workbook['Dashboard']
row_headers = ["S.No", "Domain     ", "Successful CR Count" ,"CR Done by Automation" ,"Automation %" ,"Node touch of successful CR" ,"MOP Visited for Successful CR" ,"MOP Visited %" ,"Cancelled CR" ,"Reverted CR"]
column_headers = ["RAN_Nokia", "RAN_ZTE","RAN_ERICSSON","RAN_HUAWEI"]
fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
border = Border(left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"))
bold_font = Font(bold=True)
for col, header in enumerate(row_headers, start=1):
    cell=sheet.cell(row=1, column=col, value=header)
    cell.fill = fill  # Set the background color
    cell.border = border
    cell.font = bold_font
 
for col in range(1, len(row_headers) + 1):
    column_letter = get_column_letter(col)
    max_length = max(len(str(cell.value)) for cell in sheet[column_letter])
    adjusted_width = (max_length + 2)  # Add some padding
    sheet.column_dimensions[column_letter].width = adjusted_width

for row, header in enumerate(column_headers, start=2):
 
    cell=sheet.cell(row=row, column=2, value=header)
    cell.border = border
sheet['A2'].value = 1
sheet['A2'].border = border
sheet['A3'].value = 2
sheet['A3'].border = border
sheet['A4'].value = 3
sheet['A4'].border = border
sheet['A5'].value = 4
sheet['A5'].border = border
    #NOKIA#############################################
df_nokia = df[df['Vendor']=='Nokia']
succ_cr_count = len(df_nokia[~((df_nokia['Execution Type'] == 'Reverted') | (df_nokia['Execution Type'] == 'Cancelled'))])
sheet['C2'].value = succ_cr_count
sheet['C2'].border = border
by_automation = len(df_nokia[((df_nokia['Execution Type'] == 'Auto Execution') | (df_nokia['Execution Type'] == 'Partial Auto Execution'))])
sheet['D2'].value = by_automation
sheet['D2'].border = border
aut_percentage = by_automation/succ_cr_count*100
sheet['E2'].value = int(aut_percentage)
sheet['E2'].border = border
node_touch = df_nokia[~((df_nokia['Execution Type'] == 'Reverted') | (df_nokia['Execution Type'] == 'Cancelled'))]['Planned Count'].sum()
sheet['F2'].value = node_touch
sheet['F2'].border = border
sheet['G2'].value = succ_cr_count
sheet['G2'].border = border
sheet['H2'].value = 100
sheet['H2'].border = border
sheet['I2'].value = len(df_nokia[(df_nokia['Execution Type'] == 'Cancelled')])
sheet['I2'].border = border
sheet['J2'].value = len(df_nokia[(df_nokia['Execution Type'] == 'Reverted')])
sheet['J2'].border = border
    #                ZTE    #############################################
df_zte = df[df['Vendor']=='ZTE']
succ_cr_count = len(df_zte[~((df_zte['Execution Type'] == 'Reverted') | (df_zte['Execution Type'] == 'Cancelled'))])
sheet['C3'].value = succ_cr_count
sheet['C3'].border = border
by_automation = len(df_zte[((df_zte['Execution Type'] == 'Auto Execution') | (df_zte['Execution Type'] == 'Partial Auto Execution'))])
sheet['D3'].value = by_automation
sheet['D3'].border = border
aut_percentage = by_automation/succ_cr_count*100
sheet['E3'].value = int(aut_percentage)
sheet['E3'].border = border
node_touch = df_zte[~((df_zte['Execution Type'] == 'Reverted') | (df_zte['Execution Type'] == 'Cancelled'))]['Planned Count'].sum()
sheet['F3'].value = node_touch
sheet['F3'].border = border
sheet['G3'].value = succ_cr_count
sheet['G3'].border = border
sheet['H3'].value = 100
sheet['H3'].border = border
sheet['I3'].value = len(df_zte[(df_zte['Execution Type'] == 'Cancelled')])
sheet['I3'].border = border
sheet['J3'].value = len(df_zte[(df_zte['Execution Type'] == 'Reverted')])
sheet['J3'].border = border
#Ericsson#############################################
df_Ericsson = df[(df['Vendor'] == 'Ericsson') | (df['Vendor'] == 'ERICSSON')]
succ_cr_count = len(df_Ericsson[~((df_Ericsson['Execution Type'] == 'Reverted') | (df_Ericsson['Execution Type'] == 'Cancelled'))])
sheet['C4'].value = succ_cr_count
sheet['C4'].border = border
by_automation = len(df_Ericsson[((df_Ericsson['Execution Type'] == 'Auto Execution') | (df_Ericsson['Execution Type'] == 'Partial Auto Execution'))])
sheet['D4'].value = by_automation
sheet['D4'].border = border
aut_percentage = by_automation/succ_cr_count*100
sheet['E4'].value = int(aut_percentage)
sheet['E4'].border = border
node_touch = df_Ericsson[~((df_Ericsson['Execution Type'] == 'Reverted') | (df_Ericsson['Execution Type'] == 'Cancelled'))]['Planned Count'].sum()
sheet['F4'].value = node_touch
sheet['F3'].border = border
sheet['G4'].value = succ_cr_count
sheet['G4'].border = border
sheet['H4'].value = 100
sheet['H4'].border = border
sheet['I4'].value = len(df_Ericsson[(df_Ericsson['Execution Type'] == 'Cancelled')])
sheet['I4'].border = border
sheet['J4'].value = len(df_Ericsson[(df_Ericsson['Execution Type'] == 'Reverted')])
sheet['J4'].border = border
 #Huawe#############################################
df_Huawei = df[df['Vendor']=='Huawei']
succ_cr_count = len(df_Huawei[~((df_Huawei['Execution Type'] == 'Reverted') | (df_Huawei['Execution Type'] == 'Cancelled'))])
sheet['C5'].value = succ_cr_count
sheet['C5'].border = border
by_automation = len(df_Huawei[((df_Huawei['Execution Type'] == 'Auto Execution') | (df_Huawei['Execution Type'] == 'Partial Auto Execution'))])
sheet['D5'].value = by_automation
sheet['D5'].border = border
aut_percentage = by_automation/succ_cr_count*100
sheet['E5'].value = int(aut_percentage)
sheet['E5'].border = border
node_touch = df_Huawei[~((df_Huawei['Execution Type'] == 'Reverted') | (df_Huawei['Execution Type'] == 'Cancelled'))]['Planned Count'].sum()
sheet['F5'].value = node_touch
sheet['F5'].border = border
sheet['G5'].value = succ_cr_count
sheet['G5'].border = border
sheet['H5'].value = 100
sheet['H5'].border = border
sheet['I5'].value = len(df_Huawei[(df_Huawei['Execution Type'] == 'Cancelled')])
sheet['I5'].border = border
sheet['J5'].value = len(df_Huawei[(df_Huawei['Execution Type'] == 'Reverted')])
sheet['J5'].border = border
   # Save the updated Excel file
workbook.save(os.path.join(script_directory, 'Automation data_Week-xx.xlsx'))